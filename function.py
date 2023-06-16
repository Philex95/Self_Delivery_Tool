import shutil
import openpyxl
import os
import re
import codecs
import datetime
import csv
import pandas as pd
from openpyxl.styles import Alignment, Font, colors, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from win32com.client import DispatchEx
import barcode
from barcode.writer import ImageWriter
#全局表单单元格样式
cell_alignment = Alignment(horizontal='center', vertical='center')
cell_alignmentL = Alignment(horizontal='left', vertical='center')
cell_alignmentR = Alignment(horizontal='right', vertical='center')
cell_warp = Alignment(horizontal='center', vertical='center',wrapText=True)

cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
cell_thickBorder = Border(left=Side(style='thick'), right=Side(style='thick'),
                         top=Side(style='thick'), bottom=Side(style='thick'))

cell_boldFont = Font(name='Arial Black', size=11, color=colors.BLACK, bold=True)
cell_normalFont = Font(name='Calibri', size=9, color=colors.BLACK)
cell_smallFont = Font(name='Calibri', size=7, color=colors.BLACK)
cell_normalRedFont = Font(name='Calibri', size=9, color=colors.Color(rgb='00FF0000'))

cell_pdfNorFont = Font(name='Microsoft YaHei UI', size=16,color=colors.BLACK)
cell_pdfNumFont = Font(name='Microsoft YaHei UI', size=72,color=colors.BLACK,bold=True)
cell_pdfSmallFont = Font(name='Microsoft YaHei UI', size=12,color=colors.BLACK)
cell_pdfDateFont = Font(name='Microsoft YaHei UI', size=10,color=colors.BLACK)



def getRealMaxRow(sheet):
    i=sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break
    return real_max_row

def getRealMaxColumn(sheet):
    real_max_column = sheet.max_column
    columns = [column for column in sheet.columns]
    while real_max_column>0:
        column_dict = {c.value for c in columns[real_max_column-1]}
        if column_dict=={None}:
            real_max_column = real_max_column-1
        else:
            break
    return real_max_column

def MergeCell(tarSheet , start, end, value, border, alignment, font, fill = PatternFill("solid", fgColor="FFFFFF")):
    tarSheet[start].value = tarSheet[end].value = value
    tarSheet[start].alignment = tarSheet[end].alignment = alignment
    tarSheet[start].border = tarSheet[end].border = border
    tarSheet[start].font = tarSheet[end].font = font
    tarSheet[start].fill = tarSheet[end].fill = fill

    tarSheet.merge_cells(start +':' + end)

def SetCell(tarSheet, row, col,value, border, alignment, font, fill = PatternFill("solid", fgColor="FFFFFF")):
    tarSheet.cell(row, col).value = value
    tarSheet.cell(row, col).alignment = alignment
    tarSheet.cell(row, col).border = border
    tarSheet.cell(row, col).font = font
    tarSheet.cell(row, col).fill = fill

def GetTargetCol(targetSheet, keyCode, wrongCallBack):
    targetCol = []
    foundArr = []
    realCol = getRealMaxColumn(targetSheet)
    for key in keyCode:
        isFind = False
        for col in range(1, realCol + 1):
            if targetSheet.cell(1,col).value == key:
                targetCol.append(col)
                isFind = True
                foundArr.append(isFind)
                break
        if isFind is False:
            wrongCallBack('没有找到目标列名：')
            wrongCallBack(key)
            foundArr.append(isFind)

    if False in foundArr:
        return None
    else:
        return targetCol

def LoadDataGroup(dfData, keyCode, goodDic, wrongCallBack):
    if set(keyCode) > set(dfData.columns.tolist()):
        wrongCallBack('当前表格没有目标列表中设定的列，请检查并修改信息文件夹中的目标列表文件！')
        return None

    #keyCode[3]是产品sku表头名
    skuList = dfData[keyCode[3]].tolist()
    pltList = []
    weightList = []
    for sku in skuList:
        if sku in goodDic:
            prodData = goodDic[sku]
            PLT = prodData[6]
            singleWeight = prodData[5]
            pltList.append(PLT)
            weightList.append(singleWeight)
        else:
            pltList.append('PLT_UnKnown')
            weightList.append('SingleWeight_Unknown')
            wrongCallBack("货物属性表缺乏 productcode: %s 的信息。" % sku)

    dfData.insert(20,'PLT',pltList)
    dfData.insert(21, 'singleWeight', weightList)

    tarCol = keyCode + ['PLT','singleWeight']
    dfTarget = dfData[tarCol]
    dataList = dfTarget.values.tolist()
    return dataList

def WriteDataToSheet(dataGroup, targetSheet):
    totalQTY = 0
    totalWeight = 0
    startRow = 4
    customerDic = {}
    for item in dataGroup:
        num = item[4]
        wareHouseNum = item[5]
        customerName = item[0]
        singleWeight = item[8]
        postcode = item[1]
        phone = item[2]
        sku = item[3]
        plt = item[7]

        if wareHouseNum in customerDic:
            no = customerDic[wareHouseNum]
            item.append(no)
        else:
            newNum = len(customerDic) + 1
            customerDic.update({wareHouseNum: newNum})
            item.append(newNum)
        # print('NO:' + str(item[9]) + ',PostCode:' + str(item[1]))

        SetCell(targetSheet, startRow, 1, item[9], cell_border, cell_alignment, cell_normalFont)
        SetCell(targetSheet, startRow, 9, customerName, cell_border, cell_alignment, cell_normalFont) #CostomerName
        SetCell(targetSheet, startRow, 10, postcode, cell_border, cell_alignment, cell_normalFont) #PostCode
        SetCell(targetSheet, startRow, 11, str(phone), cell_border, cell_alignment, cell_normalFont) #PhoneNum
        targetSheet.cell(startRow, 12).number_format = '0'

        SetCell(targetSheet, startRow, 2, sku, cell_border, cell_alignment, cell_normalRedFont) #ProductCode
        SetCell(targetSheet, startRow, 3, num, cell_border, cell_alignment, cell_normalRedFont) #QTY
        totalQTY += num

        if num >= 2:
            targetSheet.cell(startRow, 2).fill = PatternFill("solid", fgColor="FFD700")
            targetSheet.cell(startRow, 3).fill = PatternFill("solid", fgColor="FFD700")

        if plt != 'PLT_UnKnown':
            weight = round(float(singleWeight * num * 0.9), 2)
            totalWeight += weight
            SetCell(targetSheet, startRow, 4, str(weight), cell_border, cell_alignment, cell_normalFont)  # Weight

            SetCell(targetSheet, startRow, 5, ' ', cell_border, cell_alignment, cell_normalFont)
            SetCell(targetSheet, startRow, 6, ' ', cell_border, cell_alignment, cell_normalFont)
            SetCell(targetSheet, startRow, 7, ' ', cell_border, cell_alignment, cell_normalFont)
            SetCell(targetSheet, startRow, 8, ' ', cell_border, cell_alignment, cell_normalFont)

            row = 5
            if plt == 'PLT 1':
                row = 5
            elif plt == 'PLT 2':
                row = 6
            elif plt == 'PLT 3':
                row = 7
            elif plt == 'PLT 4':
                row = 8
            SetCell(targetSheet, startRow, row, sku, cell_border, cell_alignment, cell_normalFont)

            if num >= 2:
                targetSheet.cell(startRow, 2).fill = PatternFill("solid", fgColor="FFD700")
                targetSheet.cell(startRow, 3).fill = PatternFill("solid", fgColor="FFD700")
                targetSheet.cell(startRow, row).fill = PatternFill("solid", fgColor="FFD700")

        else:
            SetCell(targetSheet, startRow, 4, '0', cell_border, cell_alignment, cell_normalFont)  # Weight
            SetCell(targetSheet, startRow, 5, ' ', cell_border, cell_alignment, cell_normalFont, PatternFill("solid", fgColor="FA8072"))
            SetCell(targetSheet, startRow, 6, ' ', cell_border, cell_alignment, cell_normalFont, PatternFill("solid", fgColor="FA8072"))
            SetCell(targetSheet, startRow, 7, ' ', cell_border, cell_alignment, cell_normalFont, PatternFill("solid", fgColor="FA8072"))
            SetCell(targetSheet, startRow, 8, ' ', cell_border, cell_alignment, cell_normalFont, PatternFill("solid", fgColor="FA8072"))

        if num >= 2:
            targetSheet.cell(startRow, 2).fill = PatternFill("solid", fgColor="FFD700")
            targetSheet.cell(startRow, 3).fill = PatternFill("solid", fgColor="FFD700")

        startRow = startRow + 1
    data = (totalQTY, totalWeight)
    return data

def CountProCodeToSheet(dataGroup, targetSheet):

    dicPLT1 = {}
    dicPLT2 = {}
    dicPLT3 = {}
    dicPLT4 = {}

    for item in dataGroup:
        # 字典中找到货物相关信息数组
        proCode = item[3]
        num = item[4]
        if item[7] != 'PLT_UnKnown':
            if item[7] == 'PLT 1':
                UpdatePLTdic(proCode,num,dicPLT1)
            elif item[7] == 'PLT 2':
                UpdatePLTdic(proCode,num,dicPLT2)
            elif item[7] == 'PLT 3':
                UpdatePLTdic(proCode,num,dicPLT3)
            elif item[7] == 'PLT 4':
                UpdatePLTdic(proCode,num,dicPLT4)

    WritePLTdicData(targetSheet, dicPLT1, 5, 1)
    WritePLTdicData(targetSheet, dicPLT2, 5, 3)
    WritePLTdicData(targetSheet, dicPLT3, 5, 5)
    WritePLTdicData(targetSheet, dicPLT4, 5, 7)

def UpdatePLTdic(proCode,num,targetDic):
    if proCode in targetDic:
        targetDic[proCode] = targetDic[proCode] + num
    else:
        targetDic.update({proCode : num})

def WritePLTdicData(targetSheet, targetDic, targetRow ,targetCol):
    row = targetRow
    total = 0
    for item in targetDic:
        SetCell(targetSheet, row, targetCol, str(item), cell_border,cell_alignment,cell_normalFont)
        SetCell(targetSheet, row, targetCol + 1, str(targetDic[item]),cell_border,cell_alignment,cell_normalFont)
        total = total + targetDic[item]
        row = row + 1

    SetCell(targetSheet, row, targetCol, "Total", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, row, targetCol + 1, str(total), cell_border, cell_alignment, cell_normalFont)
#表头样式
def CreateHeader(targetSheet):
    #备货表1
    MergeCell(targetSheet, 'A1', 'K1', "SUNNY SHOWERS LTD", cell_border, cell_alignment, cell_boldFont)
    MergeCell(targetSheet, 'A2', 'H2', "Goods", cell_border, cell_alignment, cell_normalFont)
    MergeCell(targetSheet, 'I2', 'K2', "Customer Information", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 1, "NO.Code", cell_border, cell_alignment, cell_normalRedFont)
    SetCell(targetSheet, 3, 2, "Code", cell_border, cell_alignment, cell_normalRedFont)
    SetCell(targetSheet, 3, 3, "QTY", cell_border, cell_alignment, cell_normalRedFont)
    SetCell(targetSheet, 3, 4, "Weight", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 5, "Pallet 1", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 6, "Pallet 2", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 7, "Pallet 3", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 8, "Pallet 4", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 9, "Customer name", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 10, "Post Code", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, 3, 11, "Phone number", cell_border, cell_alignment, cell_normalFont)

    # 调整列宽
    targetSheet.column_dimensions['A'].width = 3.0
    targetSheet.column_dimensions['B'].width = 9.5
    targetSheet.column_dimensions['C'].width = 1.88
    targetSheet.column_dimensions['D'].width = 3.88
    targetSheet.column_dimensions['E'].width = 10.0
    targetSheet.column_dimensions['F'].width = 10.0
    targetSheet.column_dimensions['G'].width = 10.0
    targetSheet.column_dimensions['H'].width = 10.0
    targetSheet.column_dimensions['I'].width = 16.13
    targetSheet.column_dimensions['J'].width = 8.75
    targetSheet.column_dimensions['K'].width = 12.88

    targetSheet.column_dimensions['N'].width = 10
    targetSheet.column_dimensions['O'].width = 6.13
    targetSheet.column_dimensions['P'].width = 10
    targetSheet.column_dimensions['Q'].width = 6.13
    targetSheet.column_dimensions['R'].width = 10
    targetSheet.column_dimensions['S'].width = 6.13
    targetSheet.column_dimensions['T'].width = 10
    targetSheet.column_dimensions['U'].width = 6.13

    # 调整行高
    targetSheet.row_dimensions[1].height = 12.95
    targetSheet.row_dimensions[2].height = 12
    targetSheet.row_dimensions[3].height = 15
    targetSheet.row_dimensions[4].height = 12

def CreateCountProCodeHeader(targetSheet, runName):
    # 备货表2
    MergeCell(targetSheet, 'A1', 'H1', "SUNNY SHOWERS LTD", cell_border, cell_alignment, cell_boldFont)
    MergeCell(targetSheet, 'A2', 'H2', "Warehouse Stocking list", cell_border, cell_alignment, cell_normalFont)
    MergeCell(targetSheet, 'A3', 'B3', "Run Name:", cell_border, cell_alignment, cell_normalFont)
    MergeCell(targetSheet, 'C3', 'H3', runName, cell_border, cell_alignmentL, cell_normalFont)
    MergeCell(targetSheet, 'A4', 'B4', "Pallet 1", cell_border, cell_alignment, cell_normalFont)
    MergeCell(targetSheet, 'C4', 'D4', "Pallet 2", cell_border, cell_alignment, cell_normalFont)
    MergeCell(targetSheet, 'E4', 'F4', "Pallet 3", cell_border, cell_alignment, cell_normalFont)
    MergeCell(targetSheet, 'G4', 'H4', "Pallet 4", cell_border, cell_alignment, cell_normalFont)

    targetSheet.column_dimensions['A'].width = 10
    targetSheet.column_dimensions['B'].width = 6.13
    targetSheet.column_dimensions['C'].width = 10
    targetSheet.column_dimensions['D'].width = 6.13
    targetSheet.column_dimensions['E'].width = 10
    targetSheet.column_dimensions['F'].width = 6.13
    targetSheet.column_dimensions['G'].width = 10
    targetSheet.column_dimensions['H'].width = 6.13

def CreateFoot(targetSheet, totalQTY, totalWeight, runName):
    currentRow = getRealMaxRow(targetSheet) + 1
    SetCell(targetSheet, currentRow, 1, "", cell_border, cell_alignment, cell_normalFont)
    SetCell(targetSheet, currentRow, 2, "Total", cell_border, cell_alignment, cell_normalRedFont)
    SetCell(targetSheet, currentRow, 3, totalQTY, cell_border, cell_alignment, cell_normalRedFont)
    SetCell(targetSheet, currentRow, 4, totalWeight, cell_border, cell_alignment, cell_normalRedFont)
    MergeCell(targetSheet, 'E'+ str(currentRow), 'F' + str(currentRow), "Driver's name：", cell_border, cell_alignmentL, cell_normalFont)
    SetCell(targetSheet, currentRow, 7, "RunName", cell_border, cell_alignment, cell_normalFont)
    MergeCell(targetSheet, 'H'+ str(currentRow), 'I' + str(currentRow), runName, cell_border, cell_alignmentL, cell_normalFont)
    MergeCell(targetSheet, 'A'+ str(currentRow + 1), 'E' + str(currentRow + 1), "Warehouse team：", cell_border, cell_alignmentL, cell_normalFont)
    MergeCell(targetSheet, 'F'+ str(currentRow + 1), 'I' + str(currentRow + 1), "I CONFIRM ALL ITEMS ARE CORRECT & IN ORDER & THE VAN HAS NO FURTHER DAMAGE", cell_border, cell_warp, cell_normalFont)
    MergeCell(targetSheet, 'J'+ str(currentRow), 'J' + str(currentRow + 1), "Sign：", cell_border, cell_alignment, cell_boldFont)
    MergeCell(targetSheet, 'K'+ str(currentRow), 'K' + str(currentRow + 1), "", cell_border, cell_alignmentL, cell_boldFont)

    targetSheet.row_dimensions[currentRow].height = 30
    targetSheet.row_dimensions[currentRow + 1].height = 30

def GetGoodsInfo(targetSheet):
    realRow = getRealMaxRow(targetSheet)
    goodDic = {}
    for row in range(2, realRow + 1):
        goodSku = targetSheet.cell(row, 1).value
        goodData = []
        for col in range (2, 9):
            goodData.append(targetSheet.cell(row, col).value)
        goodDic.update({goodSku : goodData})
    return goodDic

def CreateXlsx(dataGroup, runName, savePath):
    dataWorkBook = openpyxl.Workbook()
    dataSheet = dataWorkBook.active
    dataCountProCodeSheet = dataWorkBook.create_sheet("Sheet_2")
    CreateHeader(dataSheet)
    CreateCountProCodeHeader(dataCountProCodeSheet, runName)
    data = WriteDataToSheet(dataGroup, dataSheet)
    CreateFoot(dataSheet, data[0], data[1], runName)
    CountProCodeToSheet(dataGroup, dataCountProCodeSheet)
    dataWorkBook.save(savePath + runName + '.xlsx')

barCodeFile = '.\\条形码缓存文件夹'
def CreateBarCode(prodCode):
    if '/' in prodCode:
        procode = prodCode.replace("/","#")
    else:
        procode = prodCode
    savePath = os.path.abspath(barCodeFile) + '/' + procode + '.png'
    if os.path.exists(savePath):
        return savePath
    else:
        cod = barcode.get_barcode_class('code128')
        codPic = cod(prodCode, writer=ImageWriter())
        picPath = barCodeFile + '/' + procode
        fullname = codPic.save(picPath,{'font_path':'.\\Fonts\\DejaVuSansMono.ttf','text_distance':1,'quiet_zone':2})
        return fullname

def MergeInfo(item, keyCode, targetSheet):
    SetCell(targetSheet, 2, 2, item['Run_Name'],cell_border,cell_alignment,cell_pdfNorFont)
    strNo = '%s in %s' %(item['Num'],item['QTY'])
    SetCell(targetSheet, 2, 5, strNo, cell_border, cell_alignment, cell_pdfNorFont)
    SetCell(targetSheet, 3, 5, item[keyCode[3]], cell_border, cell_alignment, cell_pdfNorFont)
    SetCell(targetSheet, 4, 5, item[keyCode[1]], cell_border, cell_alignment, cell_pdfNorFont)
    SetCell(targetSheet, 5, 5, item[keyCode[5]], cell_border, cell_warp, cell_pdfSmallFont)
    SetCell(targetSheet, 6, 5, item[keyCode[6]], cell_border, cell_warp, cell_pdfDateFont)
    SetCell(targetSheet, 3, 2, item['NO'], cell_border, cell_alignment, cell_pdfNumFont)

    prodCode = item[keyCode[3]]
    pic = CreateBarCode(prodCode)
    img = Image(pic)
    new_size = (250, 120)
    img.width, img.height = new_size
    targetSheet.add_image(img,'C8')

def CreateWordAndPDF(dataGroup, keyCode, runName, savePath, wrongCallBack):
    sortedGroup = SortedByPLT(dataGroup)
    docList = []
    num = 0
    for item in sortedGroup:
        No = item[9]
        prodCode = item[3]
        postCode = item[1]
        costomerPO = item[5]
        jobWindow = item[6]
        QTY = item[4]
        for i in range(0, QTY):
            singleDocDic = {}
            singleDocDic['NO'] = str(No)
            singleDocDic['Num'] = str(i + 1)
            singleDocDic['QTY'] = str(QTY)
            singleDocDic['Run_Name'] = runName
            singleDocDic[keyCode[3]] = prodCode
            singleDocDic[keyCode[1]] = postCode
            singleDocDic[keyCode[5]] = costomerPO
            singleDocDic[keyCode[6]] = jobWindow
            docList.append([num,singleDocDic])
            num += 1

    saveName = runName + '_PDF.xlsx'
    saveFile = savePath + saveName
    modelFile = './信息/PDF导出模板表.xlsx'

    if os.path.isfile(saveFile):
        os.remove(saveFile)
    shutil.copyfile(modelFile, saveFile)

    pdfBook = openpyxl.load_workbook(saveFile)
    page = len(docList)
    source = pdfBook.worksheets[0]
    for i in range(1, page):
        pdfSheet = pdfBook.copy_worksheet(source)
        pdfSheet.page_setup.paperHeight = '105mm'
        pdfSheet.page_setup.paperWidth = '148mm'
        pdfSheet.orientation = 'landscape'

    for item in docList:
        index = item[0]
        pdfSheet = pdfBook.worksheets[index]
        pdfSheet.title = str(index)
        MergeInfo(item[1], keyCode, pdfSheet)

    pdfBook.save(saveFile)
    pdfBook.close()

    PDFName = runName + '.pdf'
    PDFfile = savePath + PDFName
    ExcelToPDF(saveFile,PDFfile)

def ExcelToPDF(excelFile,pdfFile):
    if os.path.isfile(pdfFile):
        os.remove(pdfFile)

    xlApp = DispatchEx("Excel.Application")
    xlApp.Visible = False
    xlApp.DisplayAlerts = 0
    books = xlApp.Workbooks.Open(excelFile, False)
    books.ExportAsFixedFormat(0, pdfFile)
    books.Close(False)
    xlApp.Quit()

def ConbineWareNum(dataDic):
    groupData = []
    for item in dataDic:
        keyArr = dataDic[item]
        conbineWareNum = ''
        for item in keyArr:
            if item[5] not in conbineWareNum:
                if conbineWareNum == '':
                    conbineWareNum += item[5]
                else:
                    conbineWareNum += ' ' + item[5]

        for item in keyArr:
            item[5] = conbineWareNum
        groupData.extend(keyArr)
    return groupData

def dataGroupBy(data):
    dataDic = {}
    # 相同顾客与邮编的视为合并单,用字典累计同顾客同邮编的信息
    for item in data:
        name = item[0]
        postCode = item[1]
        sku = item[3]
        num = item[4]
        targetKey = '%s&%s' % (name,postCode)
        if targetKey in dataDic:
            targetData = dataDic[targetKey]
            isAdd = True
            for tarItem in targetData:
                if tarItem[3] == sku:
                    tarItem[4] += num
                    isAdd = False
                    break
            if isAdd:
                targetData.append(item)
        else:
            dataDic.update({targetKey:[item]})

    # 统一修改合并仓库单号
    groupData = ConbineWareNum(dataDic)
    return groupData




def SortedByPLT(dataGroup):
    sortedListPLT1 = []
    sortedListPLT2 = []
    sortedListPLT3 = []
    sortedListPLT4 = []
    sortedListPLTUnKnown = []
    sortedGroup = []

    for item in dataGroup:
        if item[7] == 'PLT 1':
            sortedListPLT1.append(item)
        elif item[7] == 'PLT 2':
            sortedListPLT2.append(item)
        elif item[7] == 'PLT 3':
            sortedListPLT3.append(item)
        elif item[7] == 'PLT 4':
            sortedListPLT4.append(item)
        else:
            sortedListPLTUnKnown.append(item)

    sortedGroup.extend(sortedListPLT1)
    sortedGroup.extend(sortedListPLT3)
    sortedGroup.extend(sortedListPLT4)
    sortedGroup.extend(sortedListPLT2)
    sortedGroup.extend(sortedListPLTUnKnown)
    return sortedGroup

def ReadDataFromFile(fileLocation, fileType):
    fileDic = {}
    dirs = os.listdir(fileLocation)  # 获取指定路径下的文件
    for file in dirs:
        if file.endswith(fileType):
            nameArr = re.split('[_#.]',file)
            realName = '%s-%s-%s-%s' %(nameArr[1],nameArr[2],nameArr[3],nameArr[4])
            fileDic.update({fileLocation + '/' + file : realName})
    return fileDic

#导出csv方法

def GetProductWeight(targetSheet):
    realRow = getRealMaxRow(targetSheet)
    goodDic = {}
    for row in range(2, realRow + 1):
        goodSku = targetSheet.cell(row, 1).value
        goodWeight = targetSheet.cell(row, 7).value
        goodDic.update({goodSku : goodWeight})
    return goodDic

def WriteData(targetSheet,data, goodDic, callBack):
    currentRow = getRealMaxRow(targetSheet) + 1
    #第一行表头
    targetSheet.cell(currentRow + 1, 1).value = 'Line Type'
    targetSheet.cell(currentRow + 1, 2).value = 'Job Type'
    targetSheet.cell(currentRow + 1, 3).value = 'PO Number'
    targetSheet.cell(currentRow + 1, 4).value = 'Job Ref'
    targetSheet.cell(currentRow + 1, 5).value = 'Run Name'
    targetSheet.cell(currentRow + 1, 6).value = 'Run Date'
    targetSheet.cell(currentRow + 1, 7).value = 'Drop Sequence'
    targetSheet.cell(currentRow + 1, 8).value = 'Start Date'
    targetSheet.cell(currentRow + 1, 9).value = 'End Date'
    targetSheet.cell(currentRow + 1, 10).value = 'Delivery Instructions'
    targetSheet.cell(currentRow + 1, 11).value = 'Customer Account Number'
    targetSheet.cell(currentRow + 1, 12).value = 'Customer Name'
    targetSheet.cell(currentRow + 1, 13).value = 'Customer Address1'
    targetSheet.cell(currentRow + 1, 14).value = 'Customer Address2'
    targetSheet.cell(currentRow + 1, 15).value = 'Customer Address3'
    targetSheet.cell(currentRow + 1, 16).value = 'Customer City'
    targetSheet.cell(currentRow + 1, 17).value = 'Customer Region'
    targetSheet.cell(currentRow + 1, 18).value = 'Customer Postcode'
    targetSheet.cell(currentRow + 1, 19).value = 'Customer Country'
    targetSheet.cell(currentRow + 1, 20).value = 'Customer Tel. No.'
    targetSheet.cell(currentRow + 1, 21).value = 'Customer Email'
    targetSheet.cell(currentRow + 1, 22).value = 'Site Name'
    targetSheet.cell(currentRow + 1, 23).value = 'Site Address1'
    targetSheet.cell(currentRow + 1, 24).value = 'Site Address2'
    targetSheet.cell(currentRow + 1, 25).value = 'Site Address3'
    targetSheet.cell(currentRow + 1, 26).value = 'Site City'
    targetSheet.cell(currentRow + 1, 27).value = 'Site Region'
    targetSheet.cell(currentRow + 1, 28).value = 'Site Postcode'
    targetSheet.cell(currentRow + 1, 29).value = 'Site Country'
    targetSheet.cell(currentRow + 1, 30).value = 'Site Latitude'
    targetSheet.cell(currentRow + 1, 31).value = 'Site Longitude'
    targetSheet.cell(currentRow + 1, 32).value = 'Site Tel. No.'
    targetSheet.cell(currentRow + 1, 33).value = 'Site Email'
    targetSheet.cell(currentRow + 1, 34).value = 'Depot'
    targetSheet.cell(currentRow + 1, 35).value = 'Driver Login'
    targetSheet.cell(currentRow + 1, 36).value = 'Drop Time'
    targetSheet.cell(currentRow + 1, 37).value = 'Internal Job Notes'
    targetSheet.cell(currentRow + 1, 38).value = 'Customer Auto Email POD'
    targetSheet.cell(currentRow + 1, 39).value = 'Site Auto Email POD'
    targetSheet.cell(currentRow + 1, 40).value = 'Site Email Notification'
    targetSheet.cell(currentRow + 1, 41).value = 'Site SMS Notification'
    targetSheet.cell(currentRow + 1, 42).value = 'Job Price'

    #第二行信息
    targetSheet.cell(currentRow + 2, 1).value = 'H'
    targetSheet.cell(currentRow + 2, 2).value = 'Delivery'

    strTracking = data[6]
    strTrackList = re.split(r'\s',strTracking)

    for item in range(0,len(strTrackList)):
        str = strTrackList[item]
        if 'AM' in str and 'E' not in str:
            strTrackList[item] = str.replace('AM', 'EAM')
        if 'SO' in str and 'E' not in str:
            strTrackList[item] = str.replace('SO', 'ESO')

    for i in range(1, len(strTrackList)):
        strTrackList.insert(i,' ')

    strTrack = ''.join(strTrackList)

    targetSheet.cell(currentRow + 2, 3).value = strTrack
    targetSheet.cell(currentRow + 2, 4).value = data[10]


    strOffect = datetime.date.today() + datetime.timedelta(days=1)
    startDate = strOffect.strftime("%Y-%m-%d 5:00")
    endOffect = datetime.date.today() + datetime.timedelta(days=4)
    endDate = endOffect.strftime("%Y-%m-%d 23:00")
    targetSheet.cell(currentRow + 2, 8).value = startDate
    targetSheet.cell(currentRow + 2, 9).value = endDate
    targetSheet.cell(currentRow + 2, 10).value = data[9]
    targetSheet.cell(currentRow + 2, 12).value = data[0]
    targetSheet.cell(currentRow + 2, 13).value = data[1]
    targetSheet.cell(currentRow + 2, 14).value = data[2]
    targetSheet.cell(currentRow + 2, 15).value = data[3]
    targetSheet.cell(currentRow + 2, 16).value = data[4]
    targetSheet.cell(currentRow + 2, 17).value = data[11]
    targetSheet.cell(currentRow + 2, 18).value = data[5]
    targetSheet.cell(currentRow + 2, 19).value = 'GB'
    targetSheet.cell(currentRow + 2, 20).value = data[9]
    targetSheet.cell(currentRow + 2, 22).value = data[0]
    targetSheet.cell(currentRow + 2, 23).value = data[1]
    targetSheet.cell(currentRow + 2, 24).value = data[2]
    targetSheet.cell(currentRow + 2, 25).value = data[3]
    targetSheet.cell(currentRow + 2, 26).value = data[4]
    targetSheet.cell(currentRow + 2, 27).value = data[11]
    targetSheet.cell(currentRow + 2, 28).value = data[5]
    targetSheet.cell(currentRow + 2, 29).value = 'GB'
    targetSheet.cell(currentRow + 2, 32).value = data[9]
    targetSheet.cell(currentRow + 2, 33).value = data[8]
    targetSheet.cell(currentRow + 2, 34).value = 'Elegant showers'
    targetSheet.cell(currentRow + 2, 35).value = 'driver1'
    targetSheet.cell(currentRow + 2, 36).value = '5'
    targetSheet.cell(currentRow + 2, 37).value = 'Call Customer before scheduling job'
    targetSheet.cell(currentRow + 2, 38).value = '1'
    targetSheet.cell(currentRow + 2, 39).value = '0'
    targetSheet.cell(currentRow + 2, 40).value = '1'
    targetSheet.cell(currentRow + 2, 41).value = '1'
    targetSheet.cell(currentRow + 2, 42).value = '5'

    #第四行表头
    targetSheet.cell(currentRow + 4, 1).value = 'Line Type'
    targetSheet.cell(currentRow + 4, 2).value = 'JobRef'
    targetSheet.cell(currentRow + 4, 3).value = 'Barcode'
    targetSheet.cell(currentRow + 4, 4).value = 'Merch Group'
    targetSheet.cell(currentRow + 4, 5).value = 'Product Code'
    targetSheet.cell(currentRow + 4, 6).value = 'Product Description'
    targetSheet.cell(currentRow + 4, 7).value = 'Quantity'
    targetSheet.cell(currentRow + 4, 8).value = 'Weight (KG)'
    targetSheet.cell(currentRow + 4, 9).value = 'Price'

    skuData = SplitSKU(data[7], goodDic,callBack)
    row = currentRow + 5
    for sku in skuData:
        targetSheet.cell(row, 1).value = 'D'
        targetSheet.cell(row, 2).value = data[10]
        targetSheet.cell(row, 3).value = sku[0]
        targetSheet.cell(row, 4).value = 'goods'
        targetSheet.cell(row, 5).value = sku[0]
        targetSheet.cell(row, 6).value = 'Furniture'
        targetSheet.cell(row, 7).value = sku[1]
        targetSheet.cell(row, 8).value = sku[2]
        if sku[2] == '0':
            str = '目标订单号：%s,sky:%s 重量为0....' %(strTrack, sku[0])
            callBack(str)
        targetSheet.cell(row, 9).value = '0'
        row += 1

def SplitSKU(data,skuDic,callBack):
    dataGroup = []
    skuList = data.split('+')
    for sku in skuList:
        skuInf = sku.split('#',2)
        weight = GetSKUWeight(skuDic,skuInf[0])
        skudata = (skuInf[0],skuInf[1], weight)
        dataGroup.append(skudata)
        if weight == '0':
            str = "货物信息缺失,prodCode：" + skuInf[0] + ",请补充信息"
            callBack(str)
    return dataGroup

def GetSKUWeight(dic,sku):
    if sku in dic:
        return str(dic[sku])
    else:
        return '0'

def CellectData(rowId, tarWorkSheet):
    DataGroup = []
    tarCol = (1,2,3,4,5,7,10,11,12,13,14,15)
    # 0 Name
    # 1 Addrline1
    # 2 Addrline2
    # 3 Addrline3
    # 4 town
    # 5 postCode
    # 6 Reference
    # 7 sku
    # 8 email
    # 9 PhoneNum
    # 10 ORDER NO.
    # 11 ZONE
    for i in tarCol:
        data = str(tarWorkSheet.cell(rowId, i).value)
        if data == 'None':
            data = '0'
        DataGroup.append(data)
    return DataGroup

def sortByZone(targetSheet):
    zoneGroup0 = []
    zoneGroup1 = []
    zoneGroup2 = []
    zoneGroup3 = []
    zoneGroup4 = []
    zoneGroup5 = []
    zoneGroup6 = []
    zoneGroup7 = []
    zoneGroup8 = []
    zoneGroup9 = []
    zoneGroup10 = []
    zoneGroup11 = []
    zoneGroup12 = []
    zoneGroup13 = []

    row = getRealMaxRow(targetSheet)
    for row in range(2, row + 1):
        data = CellectData(row, targetSheet)
        if data[11] == 'ZONE 0':
            zoneGroup0.append(data)
        elif data[11] == 'ZONE 1':
            zoneGroup1.append(data)
        elif data[11] == 'ZONE 2':
            zoneGroup2.append(data)
        elif data[11] == 'ZONE 3':
            zoneGroup3.append(data)
        elif data[11] == 'ZONE 4':
            zoneGroup4.append(data)
        elif data[11] == 'ZONE 5':
            zoneGroup5.append(data)
        elif data[11] == 'ZONE 6':
            zoneGroup6.append(data)
        elif data[11] == 'ZONE 7':
            zoneGroup7.append(data)
        elif data[11] == 'ZONE 8':
            zoneGroup8.append(data)
        elif data[11] == 'ZONE 9':
            zoneGroup9.append(data)
        elif data[11] == 'ZONE 10':
            zoneGroup10.append(data)
        elif data[11] == 'ZONE 11':
            zoneGroup11.append(data)
        elif data[11] == 'ZONE 12':
            zoneGroup12.append(data)
        elif data[11] == 'ZONE 13':
            zoneGroup13.append(data)
    dataDic = {}
    dataDic.update(
        {'ZONE 0': zoneGroup0,
        'ZONE 1': zoneGroup1,
        'ZONE 2': zoneGroup2,
        'ZONE 3': zoneGroup3,
        'ZONE 4': zoneGroup4,
        'ZONE 5': zoneGroup5,
        'ZONE 6': zoneGroup6,
        'ZONE 7': zoneGroup7,
        'ZONE 8': zoneGroup8,
        'ZONE 9': zoneGroup9,
        'ZONE 10': zoneGroup10,
        'ZONE 11': zoneGroup11,
        'ZONE 12': zoneGroup12,
        'ZONE 13': zoneGroup13
         }
    )
    return dataDic

def CreateCSV(targetPath, goodDic ,callBack, wrongCallBack):
    dataWorkBook = openpyxl.load_workbook(targetPath)
    dataSheet = dataWorkBook.worksheets[0]

    dataDic = sortByZone(dataSheet)
    zones = ('ZONE 0','ZONE 1','ZONE 2','ZONE 3','ZONE 4',
             'ZONE 5','ZONE 6','ZONE 7','ZONE 8','ZONE 9',
             'ZONE 10','ZONE 11','ZONE 12','ZONE 13')
    total = 0
    for zone in zones:
        if len(dataDic[zone]) > 0:
            callBack('正在处理 ' + zone + '区域,共' + str(len(dataDic[zone])) + '条数据。' )
            csvWorkBook = openpyxl.Workbook()
            csvSheet = csvWorkBook.active
            total += len(dataDic[zone])
            for item in dataDic[zone]:
                WriteData(csvSheet, item, goodDic, wrongCallBack)
            strOffect = datetime.date.today() + datetime.timedelta(days=1)
            dataName = "Result_" + "_" + zone + "_" +strOffect.strftime("%m-%d")
            callBack(zone + ' 处理完成..')
            csvWorkBook.save('.\\' + dataName + '.xlsx')
            Xlsx2Csv('.\\' + dataName + '.xlsx','.\\PODFather导入csv输出\\' + dataName + '.csv')
            os.remove('.\\' + dataName + '.xlsx')
        else:
            callBack(zone + '区域没有数据。')
    callBack('一共处理 ' + str(total) + '条数据...')
    callBack("处理完成!")
    return

def Xlsx2Csv(xlsxPath, SavePath):
    workbook = openpyxl.load_workbook(xlsxPath)
    workSheet = workbook.worksheets[0]
    with codecs.open(SavePath, 'w', encoding='utf-8') as f:
        write = csv.writer(f)
        currentRow = getRealMaxRow(workSheet)
        currentCol = getRealMaxColumn(workSheet)
        for row_num in range(1, currentRow + 1):
            rowData = []
            for col_num in range(1, currentCol + 1):
                rowData.append(workSheet.cell(row_num,col_num).value)
            write.writerow(rowData)

